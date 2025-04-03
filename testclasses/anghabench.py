from pathlib import Path
import sys,shutil,os,fnmatch,tarfile,subprocess,signal,resource
from openpyxl import Workbook
import asyncio,aiofiles



class AnghaBench:
    def __init__(self, **kwargs):
        self.rpms = set()
        self.believe_tmp:bool = kwargs.get('believe_tmp')
        self.path = Path('/root/osmts_tmp/AnghaBench')
        self.directory: Path = kwargs.get('saved_directory') / 'anghabench'
        self.log_files:Path = self.directory / 'log_files'
        self.matches:list = []
        self.appended_list:list = []

        self.failed = 0
        self.total = 0

        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = 'AnghaBench'

        self.ws.cell(1,1,"AnghaBench测试中编译未通过项目汇总")
        self.ws.merge_cells("A1:B1")
        self.ws.cell(2,1,"c文件名")
        self.ws.cell(2,2,"日志文件")


    def pre_test(self):
        resource.setrlimit(resource.RLIMIT_NOFILE,(65536,524288))
        if self.directory.exists():
            shutil.rmtree(self.directory)
        self.directory.mkdir(parents=True)
        self.log_files.mkdir(parents=True)

        if self.path.exists() and self.believe_tmp:
            pass
        else:
            shutil.rmtree(self.path, ignore_errors=True)
            # 拉取AnghaBench的源码
            git_clone = subprocess.run(
                f"cd /root/osmts_tmp/ && git clone https://gitcode.com/qq_61653333/AnghaBench.git",
                shell=True,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.PIPE,
            )
            if git_clone.returncode != 0:
                print(f"AnghaBench测试出错.git clone运行失败,报错信息:{git_clone.stderr.decode('utf-8')}")
                sys.exit(1)


    async def match2result(self):
        try:
            while True:
                match:list = await self.queue.get()
                log_name = match[0] + '.log'
                compile = await asyncio.create_subprocess_shell(
                    f"gcc {match[1]} -c -o {match[1]}.o",
                    stdout=asyncio.subprocess.PIPE,
                    stderr=asyncio.subprocess.STDOUT
                )
                stdout, _ = await compile.communicate()
                if compile.returncode != 0:
                    self.appended_list.append([match[0], log_name])
                    async with aiofiles.open(self.log_files / log_name, 'w') as log:
                        await log.write(stdout.decode('utf-8'))
                self.queue.task_done()
        except asyncio.CancelledError: # 取消协程
            pass


    async def run_test(self):
        if self.matches == []:
            for root,dirnames,filenames in os.walk(self.path):
                for filename in fnmatch.filter(filenames, '*.c'):
                    self.matches.append((filename,os.path.join(root,filename)))
        self.total = len(self.matches)

        self.queue = asyncio.Queue(maxsize=len(self.matches))
        for match in self.matches:
            await self.queue.put(match)
        workers = [asyncio.create_task(self.match2result()) for i in range(os.cpu_count() * 128)]

        print(f"  当前线程的event loop策略:{asyncio.get_event_loop_policy()}")

        loop = asyncio.get_event_loop()

        def signal_handler():
            print("  osmts检测到Ctrl+C键盘中断信号,正在终止AnghaBench测试...")
            for worker in workers:
                worker.cancel()
            print(f"运行至此,编译失败的数量为{len(self.appended_list)}")
            sys.exit(1)
        loop.add_signal_handler(signal.SIGINT, signal_handler)

        await self.queue.join()
        print(f"  AnghaBench测试编译结束,正在清理所有compile_worker...")
        for worker in workers:
            worker.cancel()

        for item in self.appended_list:
            self.ws.append(item)
        self.failed = len(self.appended_list)
        # 汇总结果
        self.ws.append([f"总共编译文件数{self.total}",f"通过编译数{self.total - self.failed}",f"失败编译数{self.failed}"])

        self.wb.save(self.directory / 'AnghaBench.xlsx')

        with tarfile.open(self.directory / 'AnghaBench.tar.xz',"w:xz") as tar:
            tar.add(self.log_files)


    def run(self):
        print('开始进行AnghaBench测试')
        self.pre_test()
        asyncio.run(self.run_test())
        print('AnghaBench测试结束')