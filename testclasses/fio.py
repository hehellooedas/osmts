from pathlib import Path
import sys,subprocess,shutil,hashlib,requests,re
from multiprocessing import Process
from openpyxl.workbook import Workbook



class Fio:
    def __init__(self, **kwargs):
        self.path = Path('/root/osmts_tmp/fio')
        self.directory: Path = kwargs.get('saved_directory')
        self.remove_osmts_tmp_dir:bool = kwargs.get('remove_osmts_tmp_dir')
        # 如果iso文件已经存在则不重复下载(用哈希值校验文件)
        if Path.exists(self.path / 'openEuler-24.03-LLVM-riscv64-dvd.iso'):
            iso_hash = hashlib.sha256()
            with open(self.path / 'openEuler-24.03-LLVM-riscv64-dvd.iso','rb') as file:
                while chunk := file.read(8192):
                    iso_hash.update(chunk)
            if iso_hash.hexdigest() == '74e9ac072b6b72744f21fec030fbe67ea331047ae44b26277f9d5ef41ab6776d':
                self.download_iso_process = None
                return
        self.download_iso_process:Process = Process(target=self.download_iso_file,daemon=True)
        self.download_iso_process.start()



    def download_iso_file(self):
        if self.path.exists():
            shutil.rmtree(self.path)
        self.path.mkdir()
        with requests.get(
                'https://repo.openeuler.org/openEuler-preview/openEuler-24.03-LLVM-Preview/ISO/riscv64/openEuler-24.03-LLVM-riscv64-dvd.iso',
                stream=True
        ) as response,open(self.path / 'openEuler-24.03-LLVM-riscv64-dvd.iso', 'wb') as file:
            response.raise_for_status()
            for chunk in response.iter_content(64 * 1024):
                file.write(chunk)



    def pre_test(self):
        install_fio = subprocess.run("dnf install fio -y", shell=True, stdout=subprocess.DEVNULL,
                                     stderr=subprocess.PIPE)
        if install_fio.returncode != 0:
            print(f"fio运行出错:fio包安装失败.报错信息:{install_fio.stderr.decode('utf-8')}")
            sys.exit(1)



    def run_test(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "fio"
        baseline = 1
        if self.download_iso_process is not None:
            self.download_iso_process.join()
        filename = "/root/osmts_tmp/fio/openEuler-24.03-LLVM-riscv64-dvd.iso"
        numjobs = 10
        iodepth = 10
        for rw in ("read","write","rw","randread","randwrite","randrw"):
            for bs in (4,16,32,64,128,256,512,1024):
                if rw == "randrw" or rw == "rw":
                    fio = subprocess.run(f"fio -filename={filename} -direct=1 -iodepth {iodepth} -thread -rw={rw} -rwmixread=70 -ioengine=libaio -bs={bs}k -size=1G -numjobs={numjobs} -runtime=30 -group_reporting -name={rw}-{bs}k",shell=True,stdout=subprocess.PIPE,stderr=subprocess.PIPE)
                else:
                    fio = subprocess.run(f"fio -filename={filename} -direct=1 -iodepth {iodepth} -thread -rw={rw} -ioengine=libaio -bs={bs}k -size=1G -numjobs={numjobs} -runtime=30 -group_reporting -name={rw}-{bs}k",shell=True,stdout=subprocess.PIPE,stderr=subprocess.PIPE)
                if fio.returncode != 0:
                    print(f"fio测试出错:fio进程运行报错,此时rw为{rw}.报错信息:{fio.stderr.decode('utf-8')}")
                    sys.exit(1)
                # 保存fio命令的输出结果
                result = fio.stdout.decode('utf-8')

                # 表头
                ws.cell(baseline,1,f'{rw}-{bs}k')
                ws.cell(baseline+1,2,'IOPS(k)')
                ws.cell(baseline+2,2,'BW(MiB/s)')

                # read: IOPS=47.1k, BW=184MiB/s (193MB/s)(5524MiB/30002msec)
                match = re.search(r'read: IOPS=([\d\.]+)k, BW=(\d+)MiB/s.',result)
                if match:
                    ws.cell(baseline + 1, 3, match.group(1)) # 47.1
                    ws.cell(baseline + 2, 3, match.group(2)) # 184


                # slat (usec): min=24, max=4289, avg=44.17, stdev=11.76
                ws.cell(baseline + 3, 1, 'slat (usec)')
                ws.merge_cells(f"a{baseline+3}:a{baseline+6}")
                ws.cell(baseline + 3, 2, 'min')
                ws.cell(baseline + 4, 2, 'max')
                ws.cell(baseline + 5, 2, 'avg')
                ws.cell(baseline + 6, 2, 'stdev')
                match = re.search(r"slat \(usec\): min=(\d+), max=(\d+), avg=([\d\.]+), stdev=([\d\.]+)",result)
                if match:
                    ws.cell(baseline + 3, 3, match.group(1))
                    ws.cell(baseline + 4, 3, match.group(2))
                    ws.cell(baseline + 5, 3, match.group(3))
                    ws.cell(baseline + 6, 3, match.group(4))


                # clat (usec): min=347, max=6606, avg=2073.57, stdev=266.35
                ws.cell(baseline + 7, 1, 'slat (usec)')
                ws.merge_cells(f"a{baseline+7}:a{baseline+10}")
                ws.cell(baseline + 7, 2, 'min')
                ws.cell(baseline + 8, 2, 'max')
                ws.cell(baseline + 9, 2, 'avg')
                ws.cell(baseline + 10, 2, 'stdev')
                match = re.search(r"clat \(usec\): min=(\d+), max=(\d+), avg=([\d\.]+), stdev=([\d\.]+)", result)
                if match:
                    ws.cell(baseline + 7, 3, match.group(1))
                    ws.cell(baseline + 8, 3, match.group(2))
                    ws.cell(baseline + 9, 3, match.group(3))
                    ws.cell(baseline + 10, 3, match.group(4))

                # lat (usec): min=376, max=6646, avg=2117.74, stdev=265.67
                ws.cell(baseline + 11, 1, 'slat (usec)')
                ws.merge_cells(f"a{baseline+11}:a{baseline+14}")
                ws.cell(baseline + 11, 2, 'min')
                ws.cell(baseline + 12, 2, 'max')
                ws.cell(baseline + 13, 2, 'avg')
                ws.cell(baseline + 14, 2, 'stdev')
                # 以任意个空格开头虽然紧接着lat的那一行
                match = re.search(r"^ *lat \(usec\): min=(\d+), max=(\d+), avg=([\d\.]+), stdev=([\d\.]+)", result, re.MULTILINE)
                if match:
                    ws.cell(baseline + 11, 3, match.group(1))
                    ws.cell(baseline + 12, 3, match.group(2))
                    ws.cell(baseline + 13, 3, match.group(3))
                    ws.cell(baseline + 14, 3, match.group(4))


                ws.cell(baseline + 15, 1, 'clat percentiles (usec)')
                ws.merge_cells(f"a{baseline+15}:a{baseline+31}")
                # |  1.00th=[ 1205],  5.00th=[ 1696], 10.00th=[ 1811], 20.00th=[ 1909],
                ws.cell(baseline + 15, 2, '1.00th')
                ws.cell(baseline + 16, 2, '5.00th')
                ws.cell(baseline + 17, 2, '10.00th')
                ws.cell(baseline + 18, 2, '20.00th')
                ws.cell(baseline + 19, 2, '30.00th')
                ws.cell(baseline + 20, 2, '40.00th')
                ws.cell(baseline + 21, 2, '50.00th')
                ws.cell(baseline + 22, 2, '60.00th')
                ws.cell(baseline + 23, 2, '70.00th')
                ws.cell(baseline + 24, 2, '80.00th')
                ws.cell(baseline + 25, 2, '90.00th')
                ws.cell(baseline + 26, 2, '95.00th')
                ws.cell(baseline + 27, 2, '99.00th')
                ws.cell(baseline + 28, 2, '99.50th')
                ws.cell(baseline + 29, 2, '99.90th')
                ws.cell(baseline + 30, 2, '99.95th')
                ws.cell(baseline + 31, 2, '99.99th')
                match = re.search(r'\|  1\.00th=\[ (\d+)\],  5\.00th=\[ (\d+)\], 10\.00th=\[ (\d+)\], 20\.00th=\[ (\d+)\],',result)
                if match:
                    ws.cell(baseline + 15, 3, match.group(1))
                    ws.cell(baseline + 16, 3, match.group(2))
                    ws.cell(baseline + 17, 3, match.group(3))
                    ws.cell(baseline + 18, 3, match.group(4))
                # | 30.00th=[ 1958], 40.00th=[ 2008], 50.00th=[ 2057], 60.00th=[ 2114],
                match = re.search(r'\| 30\.00th=\[ (\d+)\], 40\.00th=\[ (\d+)\], 50\.00th=\[ (\d+)\], 60\.00th=\[ (\d+)\],',result)
                if match:
                    ws.cell(baseline + 19, 3, match.group(1))
                    ws.cell(baseline + 20, 3, match.group(2))
                    ws.cell(baseline + 21, 3, match.group(3))
                    ws.cell(baseline + 22, 3, match.group(4))
                # | 70.00th=[ 2180], 80.00th=[ 2245], 90.00th=[ 2409], 95.00th=[ 2507],
                match = re.search(r'\| 70\.00th=\[ (\d+)\], 80\.00th=\[ (\d+)\], 90\.00th=\[ (\d+)\], 95\.00th=\[ (\d+)\],',result)
                if match:
                    ws.cell(baseline + 23, 3, match.group(1))
                    ws.cell(baseline + 24, 3, match.group(2))
                    ws.cell(baseline + 25, 3, match.group(3))
                    ws.cell(baseline + 26, 3, match.group(4))
                # | 99.00th=[ 2737], 99.50th=[ 2835], 99.90th=[ 2999], 99.95th=[ 3064],
                match = re.search(r'\| 99\.00th=\[ (\d+)\], 99\.50th=\[ (\d+)\], 99\.90th=\[ (\d+)\], 99\.95th=\[ (\d+)\],',result)
                if match:
                    ws.cell(baseline + 27, 3, match.group(1))
                    ws.cell(baseline + 28, 3, match.group(2))
                    ws.cell(baseline + 29, 3, match.group(3))
                    ws.cell(baseline + 30, 3, match.group(4))
                # | 99.99th=[ 3261]
                match = re.search(r'\| 99\.99th=\[ (\d+)\]',result)
                if match:
                    ws.cell(baseline + 31, 3, match.group(1))


                # bw (  KiB/s): min=165493, max=225392, per=100.00%, avg=188882.14, stdev=1195.30, samples=590
                ws.cell(baseline + 32,1,'bw (KiB/s)')
                ws.merge_cells(f"a{baseline+32}:a{baseline+37}")
                ws.cell(baseline + 32, 2, 'min')
                ws.cell(baseline + 33, 2, 'max')
                ws.cell(baseline + 34, 2, 'per')
                ws.cell(baseline + 35, 2, 'avg')
                ws.cell(baseline + 36, 2, 'stdev')
                ws.cell(baseline + 37, 2, 'samples')

                match = re.search(r'bw \(  KiB/s\): min=(\d+), max=(\d+), per=([\d\.]+)%, avg=([\d\.]+), stdev=([\d\.]+), samples=(\d+)',result)
                if match:
                    ws.cell(baseline + 32, 3, match.group(1))
                    ws.cell(baseline + 33, 3, match.group(2))
                    ws.cell(baseline + 34, 3, match.group(3))
                    ws.cell(baseline + 35, 3, match.group(4))
                    ws.cell(baseline + 36, 3, match.group(5))
                    ws.cell(baseline + 37, 3, match.group(6))

                # iops        : min=41372, max=56348, avg=47220.12, stdev=298.84, samples=590
                ws.cell(baseline + 38, 1, 'iops')
                ws.merge_cells(f"a{baseline+38}:a{baseline+42}")
                ws.cell(baseline + 38, 2, 'min')
                ws.cell(baseline + 39, 2, 'max')
                ws.cell(baseline + 40, 2, 'avg')
                ws.cell(baseline + 41, 2, 'stdev')
                ws.cell(baseline + 42, 2, 'samples')
                match = re.search(r'iops        : min=(\d+), max=(\d+), avg=([\d\.]+), stdev=([\d\.]+), samples=(\d+)',result)
                if match:
                    ws.cell(baseline + 38, 3, match.group(1))
                    ws.cell(baseline + 39, 3, match.group(2))
                    ws.cell(baseline + 40, 3, match.group(3))
                    ws.cell(baseline + 41, 3, match.group(4))
                    ws.cell(baseline + 42, 3, match.group(5))

                # cpu          : usr=3.40%, sys=23.42%, ctx=901992, majf=0, minf=1104
                ws.cell(baseline + 43, 1, 'cpu')
                ws.merge_cells(f"a{baseline+43}:a{baseline+47}")
                ws.cell(baseline + 43, 2, 'usr')
                ws.cell(baseline + 44, 2, 'sys')
                ws.cell(baseline + 45, 2, 'cts')
                ws.cell(baseline + 46, 2, 'majf')
                ws.cell(baseline + 47, 2, 'minf')
                match = re.search(r'cpu          : usr=([\d\.]+)%, sys=([\d\.]+)%, ctx=(\d+), majf=(\d+), minf=(\d+)',result)
                if match:
                    ws.cell(baseline + 43, 3, match.group(1) + '%')
                    ws.cell(baseline + 44, 3, match.group(2) + '%')
                    ws.cell(baseline + 45, 3, match.group(3))
                    ws.cell(baseline + 46, 3, match.group(4))
                    ws.cell(baseline + 47, 3, match.group(5))

                # IO depths    : 1=0.1%, 2=0.1%, 4=0.1%, 8=100.0%, 16=0.0%, 32=0.0%, >=64=0.0%
                ws.cell(baseline + 48, 1, 'IO depths')
                ws.merge_cells(f"a{baseline+48}:a{baseline+54}")
                ws.cell(baseline + 48, 2, '1')
                ws.cell(baseline + 49, 2, '2')
                ws.cell(baseline + 50, 2, '4')
                ws.cell(baseline + 51, 2, '8')
                ws.cell(baseline + 52, 2, '16')
                ws.cell(baseline + 53, 2, '32')
                ws.cell(baseline + 54, 2, '>=64')
                match = re.search(r'IO depths    : 1=([\d\.]+)%, 2=([\d\.]+)%, 4=([\d\.]+)%, 8=([\d\.]+)%, 16=([\d\.]+)%, 32=([\d\.]+)%, >=64=([\d\.]+)%',result)
                if match:
                    ws.cell(baseline + 48, 3, match.group(1) + '%')
                    ws.cell(baseline + 49, 3, match.group(2) + '%')
                    ws.cell(baseline + 50, 3, match.group(3) + '%')
                    ws.cell(baseline + 51, 3, match.group(4) + '%')
                    ws.cell(baseline + 52, 3, match.group(5) + '%')
                    ws.cell(baseline + 53, 3, match.group(6) + '%')
                    ws.cell(baseline + 54, 3, match.group(7) + '%')

                # submit    : 0=0.0%, 4=100.0%, 8=0.0%, 16=0.0%, 32=0.0%, 64=0.0%, >=64=0.0%
                ws.cell(baseline + 55, 1, 'submit')
                ws.merge_cells(f"a{baseline+55}:a{baseline+61}")
                ws.cell(baseline + 55, 2, '0')
                ws.cell(baseline + 56, 2, '4')
                ws.cell(baseline + 57, 2, '8')
                ws.cell(baseline + 58, 2, '16')
                ws.cell(baseline + 59, 2, '32')
                ws.cell(baseline + 60, 2, '64')
                ws.cell(baseline + 61, 2, '>=64')
                match = re.search(r'submit    : 0=([\d\.]+)%, 4=([\d\.]+)%, 8=([\d\.]+)%, 16=([\d\.]+)%, 32=([\d\.]+)%, 64=([\d\.]+)%, >=64=([\d\.]+)%',result)
                if match:
                    ws.cell(baseline + 55, 3, match.group(1) + '%')
                    ws.cell(baseline + 56, 3, match.group(2) + '%')
                    ws.cell(baseline + 57, 3, match.group(3) + '%')
                    ws.cell(baseline + 58, 3, match.group(4) + '%')
                    ws.cell(baseline + 59, 3, match.group(5) + '%')
                    ws.cell(baseline + 60, 3, match.group(6) + '%')
                    ws.cell(baseline + 61, 3, match.group(7) + '%')

                # complete  : 0=0.0%, 4=100.0%, 8=0.0%, 16=0.1%, 32=0.0%, 64=0.0%, >=64=0.0%
                ws.cell(baseline + 62, 1, 'complete')
                ws.merge_cells(f"a{baseline+62}:a{baseline+68}")
                ws.cell(baseline + 62, 2, '0')
                ws.cell(baseline + 63, 2, '4')
                ws.cell(baseline + 64, 2, '8')
                ws.cell(baseline + 65, 2, '16')
                ws.cell(baseline + 66, 2, '32')
                ws.cell(baseline + 67, 2, '64')
                ws.cell(baseline + 68, 2, '>=64')
                match = re.search(r'complete  : 0=([\d\.]+)%, 4=([\d\.]+)%, 8=([\d\.]+)%, 16=([\d\.]+)%, 32=([\d\.]+)%, 64=([\d\.]+)%, >=64=([\d\.]+)%',result)
                if match:
                    ws.cell(baseline + 62, 3, match.group(1) + '%')
                    ws.cell(baseline + 63, 3, match.group(2) + '%')
                    ws.cell(baseline + 64, 3, match.group(3) + '%')
                    ws.cell(baseline + 65, 3, match.group(4) + '%')
                    ws.cell(baseline + 66, 3, match.group(5) + '%')
                    ws.cell(baseline + 67, 3, match.group(6) + '%')
                    ws.cell(baseline + 68, 3, match.group(7) + '%')

                # issued rwts: total=1414061,0,0,0 short=0,0,0,0 dropped=0,0,0,0
                ws.cell(baseline + 69, 1, 'issued rwts')
                ws.merge_cells(f"a{baseline+69}:a{baseline+71}")
                ws.cell(baseline + 69, 2, 'total')
                ws.cell(baseline + 70, 2, 'short')
                ws.cell(baseline + 71, 2, 'dropped')
                match = re.search(r'issued rwts: total=([\d,]+) short=([\d,]+) dropped=([\d,]+)',result)
                if match:
                    ws.cell(baseline + 69, 3, match.group(1))
                    ws.cell(baseline + 70, 3, match.group(2))
                    ws.cell(baseline + 71, 3, match.group(3))


                # latency   : target=0, window=0, percentile=100.00%, depth=10
                ws.cell(baseline + 72, 1, 'latency')
                ws.merge_cells(f"a{baseline+72}:a{baseline+75}")
                ws.cell(baseline + 72, 2, 'target')
                ws.cell(baseline + 73, 2, 'window')
                ws.cell(baseline + 74, 2, 'percentile')
                ws.cell(baseline + 75, 2, 'depth')
                match = re.search(r'latency   : target=(\d+), window=(\d+), percentile=([\d\.]+)%, depth=(\d+)',result)
                if match:
                    ws.cell(baseline + 72, 3, match.group(1))
                    ws.cell(baseline + 73, 3, match.group(2))
                    ws.cell(baseline + 74, 3, match.group(3) + '%')
                    ws.cell(baseline + 75, 3, match.group(4))

                # READ: bw=184MiB/s (193MB/s), 184MiB/s-184MiB/s (193MB/s-193MB/s), io=5524MiB (5792MB), run=30002-30002msec
                ws.cell(baseline + 76, 1, 'Run status group 0 (all jobs)')
                ws.cell(baseline + 76, 2, 'bw(MiB/s)')
                ws.cell(baseline + 77, 2, 'io(MiB)')
                ws.cell(baseline + 78, 2, 'run(mesc)')
                match = re.search(r'^(.*?): bw=(\d+)MiB/s.*io=(\d+)MiB.*run=(\d+)-(\d+)msec', result,re.MULTILINE)
                if match:
                    ws.cell(baseline + 76, 1, 'Run status group 0 (all jobs)' + match.group(1))
                    ws.merge_cells(f"a{baseline+76}:a{baseline+78}")
                    ws.cell(baseline + 76, 3, match.group(2))
                    ws.cell(baseline + 77, 3, match.group(3))
                    ws.cell(baseline + 78, 3, match.group(4) + '-' + match.group(5))

                # nvme0n1: ios=1408818/17, merge=0/3, ticks=2897302/37, in_queue=2897344, util=81.64%
                ws.cell(baseline + 79, 1, 'Disk stats (read/write)')
                ws.merge_cells(f"a{baseline+79}:a{baseline+84}")
                ws.cell(baseline + 79, 2, 'disk name')
                ws.cell(baseline + 80, 2, 'ios')
                ws.cell(baseline + 81, 2, 'merge')
                ws.cell(baseline + 82, 2, 'ticks')
                ws.cell(baseline + 83, 2, 'in_queue')
                ws.cell(baseline + 84, 2, 'util')
                match = re.search(r'(\S+): ios=(\d+)/(\d+), merge=(\d+)/(\d+), ticks=(\d+)/(\d+), in_queue=(\d+), util=([0-9.]+%)', result)
                if match:
                    ws.cell(baseline + 79, 3, match.group(1))
                    ws.cell(baseline + 80, 3, f"{match.group(2)}/{match.group(3)}")
                    ws.cell(baseline + 81, 3, f'{match.group(4)}/{match.group(5)}')
                    ws.cell(baseline + 82, 3, f'{match.group(6)}/{match.group(7)}')
                    ws.cell(baseline + 83, 3, match.group(8))
                    ws.cell(baseline + 84, 3, match.group(9) + '%')

                baseline += 85


        wb.save(self.directory / 'fio.xlsx')


    def post_test(self):
        if self.path.exists():
            shutil.rmtree(self.path)


    def run(self):
        print("开始进行fio测试")
        self.pre_test()
        self.run_test()
        if self.remove_osmts_tmp_dir:
            self.post_test()
        print("fio测试结束")