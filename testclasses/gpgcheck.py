import shutil
import subprocess
import sys,os
import asyncio
import numpy
from pathlib import Path
from openpyxl import Workbook



class GpgCheck:
    def __init__(self, **kwargs):
        self.rpms = set()
        self.path = Path('/root/osmts_tmp/gpgcheck')
        self.directory: Path = kwargs.get('saved_directory') / 'gpgcheck'
        self.packages = []

        # 创建Excel表格
        self.wb = Workbook()
        self.ws = self.wb.active


    async def rpm_check_each(self,package_name):
        rpm_check = await asyncio.create_subprocess_shell(
            f"rpm -K /root/osmts_tmp/gpgcheck/{package_name}",
            stdout=asyncio.subprocess.DEVNULL,
            stderr=asyncio.subprocess.PIPE,
        )
        await asyncio.sleep(1.5)
        _, stderr = await rpm_check.communicate()
        if rpm_check.returncode != 0: # rpm -K运行失败
            self.ws.append([package_name, stderr.decode('utf-8')])


    async def rpm_check_all(self):
        packages = list(os.walk(self.path))[0][2]
        # 对每个rpm包创建一个测试任务
        tasks = [asyncio.create_task(self.rpm_check_each(package)) for package in packages]
        await asyncio.gather(*tasks)


    def pre_test(self):
        if not self.directory.exists():
            self.directory.mkdir()
        if self.path.exists():
            shutil.rmtree(self.path)
        self.path.mkdir(parents=True)

        # 更新缓存以便后面下载
        subprocess.run(
            "dnf clean all && dnf makecache",
            shell=True,
            stdout=subprocess.DEVNULL,stderr=subprocess.DEVNULL,
        )

        # 引入openEuler的gpg验证密钥
        import_gpg_key = subprocess.run(
            "rpm --import /etc/pki/rpm-gpg/RPM-GPG-KEY-openEuler",
            shell=True,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.PIPE,
        )
        if import_gpg_key.returncode != 0:
            print(f"gpgcheck测试出错.import gpg文件失败,报错信息:{import_gpg_key.stderr.decode('utf-8')}")
            sys.exit(1)

        # 获取已安装的所有rpm包名
        dnf_list = subprocess.run(
            "dnf list available | awk '/Available Packages/{flag=1; next} flag' | awk '{print $1}'",
            shell=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
        )
        if dnf_list.returncode != 0:
            print(f"gpgcheck测试出错.获取所有已安装的rpm包名失败,报错信息:{dnf_list.stderr.decode('utf-8')}")
            sys.exit(1)
        self.rpm_package_list = dnf_list.stdout.decode('utf-8').splitlines()

        self.ws.title = 'gpgcheck'
        self.ws.cell(1,1,f"当前系统已安装rpm包的数量:{len(self.rpm_package_list)}")
        self.ws.merge_cells('A1:B1')
        self.ws.cell(2,1,"gpgcheck失败的rpm包统计")
        self.ws.merge_cells('A2:B2')
        self.ws.cell(3,1,"package name")
        self.ws.cell(3,2,"报错日志")


    def run_test(self):
        # 排除掉不符合测试要求的包名
        for package in self.rpm_package_list:
            if package.endswith('.src') or 'debug' in package:
                continue
            else:
                self.packages.append(package)

        # 根据包名批量下载并测试rpm包
        for package_list in numpy.array_split(self.packages,200):
            rpm_download = subprocess.run(
                f"dnf download {' '.join(package_list)} --destdir={self.path}",
                shell=True,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.PIPE,
            )
            if rpm_download.returncode != 0:
                print(f"gpgcheck测试出错.下载待测的rpm包失败,报错信息:{rpm_download.stderr.decode('utf-8')}")
                print(f"本批次下载出错的rpm包为:{package_list}")

            asyncio.run(self.rpm_check_all())
            shutil.rmtree(self.path)
            self.path.mkdir(parents=True)
        self.wb.save(self.directory / 'gpgcheck.xlsx')


    def run(self):
        print('开始进行gpgcheck测试')
        self.pre_test()
        self.run_test()
        print('gpgcheck测试结束')