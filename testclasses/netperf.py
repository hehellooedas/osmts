from pathlib import Path
import re,sys,subprocess,psutil
from openpyxl import Workbook



class Netperf(object):
    def __init__(self,**kwargs):
        self.rpms = {'netperf'}
        self.directory:Path = kwargs.get('saved_directory') / 'netperf'
        self.server_ip:str = kwargs.get('netperf_server_ip')
        self.netserver_created_by_osmts:bool = kwargs.get('netserver_created_by_osmts')


    def run_test(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "netperf"

        ws.cell(1,1,"TCP STREAM TEST")
        ws.merge_cells("a1:a5")

        ws.cell(6,1,"UDP STREAM TEST")
        ws.merge_cells("a6:a18")

        ws.cell(20,1,"TCP REQUEST/RESPONSE TEST")
        ws.merge_cells("a20:a21")

        ws.cell(22,1,"TCP Connect/Request/Response TEST")
        ws.merge_cells("a22:a23")

        ws.cell(24,1,"UDP REQUEST/RESPONSE TEST")
        ws.merge_cells("a24:a25")

        # TCP_STREAM表头
        ws.cell(1,2,"Recv Socket Size bytes")
        ws.cell(1,3,"Send Socket Size Bytes")
        ws.cell(1,4,"Send Message Size Bytes")
        ws.cell(1,5,"Elapsed Time secs.")
        ws.cell(1,6,"Throughput(10^6bits/sec)")

        # UDP_STREAM表头
        ws.cell(6,2,"Socket Size bytes")
        ws.cell(6,3,"Message Size bytes")
        ws.cell(6,4,"Elapsed Time secs")
        ws.cell(6,5,"Messages Okay")
        ws.cell(6, 6, "Messages Errors")
        ws.cell(6,7,"Throughput(10^6bits/sec)")

        # 剩余三个测试的表头
        ws.cell(19,2,"Local Socket Send bytes")
        ws.cell(19, 3, "Remote Size Recv Bytes")
        ws.cell(19, 4, "Request Size bytes")
        ws.cell(19, 5, "Resp. Size bytes")
        ws.cell(19, 6, "Elapsed Time secs.")
        ws.cell(19,7,"Trans. Rate per sec")




        # TCP_STREAM测试
        line = 2
        for message_size_bytes in (1,64,512,65536):
            TCP_STREAM = subprocess.run(f"netperf -t TCP_STREAM -H {self.server_ip} -p 10000 -l 60 -- -m {message_size_bytes}",shell=True,stdout=subprocess.PIPE,stderr=subprocess.PIPE)
            if TCP_STREAM.returncode != 0:
                print(f"netperf测试出错:TCP_STREAM测试运行失败.报错信息:{TCP_STREAM.stderr}")
                sys.exit(1)
            result = re.findall(r'\d+\.\d+|\d+', TCP_STREAM.stdout.decode('utf-8').split('\n')[6])
            for col,value in enumerate(result):
                ws.cell(line,col+2,value)
            line += 1


        # UDP_STREAM测试
        line = 7
        for message_size_bytes in (1,64,128,256,512,32768):
            UDP_STREAM = subprocess.run(f"netperf -t UDP_STREAM -H {self.server_ip} -p 10000 -l 60 -- -m {message_size_bytes}",shell=True,stdout=subprocess.PIPE,stderr=subprocess.PIPE)
            if UDP_STREAM.returncode != 0:
                print(f"netperf测试出错:UDP_STREAM测试运行失败.报错信息:{UDP_STREAM.stderr.decode('utf-8')}")
                sys.exit(1)
            result1 = re.findall(r'\d+\.\d+|\d+', UDP_STREAM.stdout.decode('utf-8').split('\n')[5])
            result2 = re.findall(r'\d+\.\d+|\d+', UDP_STREAM.stdout.decode('utf-8').split('\n')[6])
            for col,value in enumerate(result1):
                ws.cell(line,col+2,value)
            line += 1
            ws.cell(line,2,result2[0])
            ws.cell(line, 4, result2[1])
            ws.cell(line, 5, result2[2])
            ws.cell(line, 7, result2[3])
            line += 1


        # TCP REQUEST/RESPONSE测试
        TCP_RR = subprocess.run(f"netperf -t TCP_RR -H {self.server_ip} -p 10000",shell=True,stdout=subprocess.PIPE,stderr=subprocess.PIPE)
        if TCP_RR.returncode != 0:
            print(f"netperf测试出错:TCP_RR测试运行失败.报错信息:{TCP_RR.stderr.decode('utf-8')}")
            sys.exit(1)
        result1 = re.findall(r'\d+\.\d+|\d+', TCP_RR.stdout.decode('utf-8').split('\n')[6])
        result2 = re.findall(r'\d+\.\d+|\d+', TCP_RR.stdout.decode('utf-8').split('\n')[7])
        for col, value in enumerate(result1):
            ws.cell(20, col + 2, value)
        ws.cell(21, 2, result2[0])
        ws.cell(21, 3, result2[1])


        # TCP Connect/Request/Response测试
        TCP_CRR = subprocess.run(f"netperf -t TCP_CRR -H {self.server_ip} -p 10000",shell=True,stdout=subprocess.PIPE,stderr=subprocess.PIPE)
        if TCP_CRR.returncode != 0:
            print(f"netperf测试出错:TCP_CRR测试运行失败.报错信息:{TCP_CRR.stderr.decode('utf-8')}")
            sys.exit(1)
        result1 = re.findall(r'\d+\.\d+|\d+', TCP_CRR.stdout.decode('utf-8').split('\n')[6])
        result2 = re.findall(r'\d+\.\d+|\d+', TCP_CRR.stdout.decode('utf-8').split('\n')[7])
        for col, value in enumerate(result1):
            ws.cell(22, col + 2, value)
        ws.cell(23, 2, result2[0])
        ws.cell(23, 3, result2[1])


        # UDP REQUEST/RESPONSE测试
        UDP_RR = subprocess.run(f"netperf -t UDP_RR -H {self.server_ip} -p 10000",shell=True,stdout=subprocess.PIPE,stderr=subprocess.PIPE)
        if UDP_RR.returncode != 0:
            print(f"netperf测试出错:UDP_RR测试运行失败.报错信息:{UDP_RR.stderr.decode('utf-8')}")
            sys.exit(1)
        result1 = re.findall(r'\d+\.\d+|\d+', UDP_RR.stdout.decode('utf-8').split('\n')[6])
        result2 = re.findall(r'\d+\.\d+|\d+', UDP_RR.stdout.decode('utf-8').split('\n')[7])
        for col, value in enumerate(result1):
            ws.cell(24, col + 2, value)
        ws.cell(25, 2, result2[0])
        ws.cell(25, 3, result2[1])

        if not self.directory.exists():
            self.directory.mkdir(exist_ok=True,parents=True)
        wb.save(self.directory / 'netperf.xlsx')


    def post_test(self):
        if self.netserver_created_by_osmts:
            for process in psutil.process_iter():
                if process.name() == 'netserver':
                    process.terminate()


    def run(self):
        print("开始进行netperf测试")
        self.run_test()
        self.post_test()
        print("netperf测试结束")