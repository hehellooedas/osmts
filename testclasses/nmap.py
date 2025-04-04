import subprocess,re
import sys
from pathlib import Path
from openpyxl import Workbook


"""
nmap测试:具体请查看文档https://gitee.com/jean9823/openEuler_riscv_test/blob/master/%E5%9C%A8openEuler%20RISC-V%2024.03%20LTS%20%E4%B8%8A%E6%89%8B%E5%8A%A8%E6%89%A7%E8%A1%8C%E6%80%A7%E8%83%BD%E6%B5%8B%E8%AF%95.md#2-nmap
"""

class Nmap:
    def __init__(self, **kwargs):
        self.rpms = {'nmap'}
        self.directory:Path = kwargs.get('saved_directory') / 'nmap'
        self.test_result = ''


    def run_test(self):
        nmap = subprocess.run(
            "nmap -sS -sU 127.0.0.1",
            shell=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE
        )
        if nmap.returncode != 0:
            print(f"nmap测试出错:nmap进程运行出错.报错信息:{nmap.stderr.decode('utf-8')}")
            sys.exit(1)
        self.test_result = nmap.stdout.decode('utf-8')


    def result2summary(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "nmap"
        ws.cell(1,1,"PORT")
        ws.cell(1,2,"STATE")
        ws.cell(1,3,"SERVICE")
        index = 2
        for port,protocol,state,service in re.findall(r"(\d+)\/(tcp|udp)\s+([\w\|\-]+)\s+(\S+)", self.test_result, re.MULTILINE | re.IGNORECASE):
            ws.cell(index,1,port+'/'+protocol)
            ws.cell(index,2,state)
            ws.cell(index,3,service)
            index += 1

        if not self.directory.exists():
            self.directory.mkdir(exist_ok=True,parents=True)
        wb.save(self.directory / 'nmap.xlsx')



    def run(self):
        print("开始进行nmap测试")
        self.run_test()
        self.result2summary()
        print("nmap测试结束")