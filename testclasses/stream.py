from pathlib import Path
import re,sys,subprocess,shutil
from openpyxl import Workbook



class Stream:
    def __init__(self,**kwargs ):
        self.rpms = set()
        self.believe_tmp: bool = kwargs.get('believe_tmp')
        self.path = Path('/root/osmts_tmp/stream')
        self.directory:Path = kwargs.get('saved_directory') / 'stream'
        self.compiler:str = kwargs.get('compiler')
        self.test_result = ''


    def pre_test(self):
        if not self.directory.exists():
            self.directory.mkdir(exist_ok=True,parents=True)

        # believe_tmp表明尽可能相信/root/osmts_tmp/
        if self.path.exists() and self.believe_tmp:
            pass
        else:
            shutil.rmtree(self.path, ignore_errors=True)
            git_clone = subprocess.run("cd /root/osmts_tmp/ && git clone https://gitee.com/April_Zhao/stream.git",shell=True,stdout=subprocess.DEVNULL,stderr=subprocess.PIPE)
            if git_clone.returncode != 0:
                print(f"stream测试出错:git拉取stream失败.报错信息:{git_clone.stderr.decode('utf-8')}")
                sys.exit(1)

        # 编译stream.c
        compile = subprocess.run(f"cd /root/osmts_tmp/stream && {self.compiler} -O3 -fopenmp -DSTREAM_ARRAY_SIZE=35000000 -DNTIMES=50 stream.c -o stream_o3",shell=True,stdout=subprocess.DEVNULL,stderr=subprocess.PIPE)
        if compile.returncode != 0:
            print(f"stream测试出错:编译stream.c未通过.报错信息:{compile.stderr.decode('utf-8')}")
            sys.exit(1)


    def run_test(self):
        stream_o3 = subprocess.run(f"cd /root/osmts_tmp/stream && ./stream_o3",shell=True,stdout=subprocess.PIPE,stderr=subprocess.PIPE)
        if stream_o3.returncode != 0:
            print(f"stream测试出错:运行./stream_o3失败.报错信息:{stream_o3.stderr.decode('utf-8')}")
        self.test_result = stream_o3.stdout.decode('utf-8')
        with open(self.path / 'stream.log','w',encoding='utf-8') as file:
            file.write(self.test_result)


    def result2summary(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "stream"
        ws.cell(1, 1, "Function")
        ws.cell(1, 2, "Best Rate MB/s")
        ws.cell(1, 3, "Avg time")
        ws.cell(1, 4, "Min time")
        ws.cell(1, 5, "Max time")
        ws.cell(2, 1, "Copy")
        ws.cell(3, 1, "Scale")
        ws.cell(4, 1, "Add")
        ws.cell(5, 1, "Triad")

        col_list = ["Copy:", "Scale:", "Add:", "Triad:"]
        for i in range(2, 6):
            Function = re.search(rf'^\s*{col_list[i - 2]}\s+([\d\.]+)\s+([\d\.]+)\s+([\d\.]+)\s+([\d\.]+)',
                                 self.test_result,
                                 re.MULTILINE | re.IGNORECASE)
            ws.cell(i, 1, col_list[i - 2])
            for j in range(2, 6):
                ws.cell(i, j, Function.group(j - 1))
        # 保存
        wb.save(self.directory / 'stream.xlsx')



    def run(self):
        print("开始进行stream测试")
        self.pre_test()
        self.run_test()
        self.result2summary()
        print("stream测试结束")