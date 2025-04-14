import subprocess,fileinput,shutil,re,sys
from pathlib import Path
from openpyxl import Workbook

from errors import GitCloneError,CompileError,RunError,SummaryError


class Unixbench:
    def __init__(self,**kwargs ):
        self.rpms = {'perl','perl-CPAN'}
        self.believe_tmp: bool = kwargs.get('believe_tmp')
        self.path = Path('/root/osmts_tmp/unixbench')
        self.directory:Path = kwargs.get('saved_directory') / 'unixbench'
        self.compiler:str = kwargs.get('compiler')
        self.test_result = ''


    def pre_test(self):
        if not self.directory.exists():
            self.directory.mkdir(exist_ok=True,parents=True)
        if self.path.exists() and self.believe_tmp:
            pass
        else:
            shutil.rmtree(self.path, ignore_errors=True)
            try:
                subprocess.run(
                    args="git clone https://gitcode.com/gh_mirrors/by/byte-unixbench.git",
                    cwd="/root/osmts_tmp",
                    shell=True,check=True,
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.PIPE
                )
            except subprocess.CalledProcessError as e:
                raise GitCloneError(e.returncode,'https://gitcode.com/gh_mirrors/by/byte-unixbench.git',e.stderr)


        if self.compiler == 'clang':
            for line in fileinput.input('/root/osmts_tmp/byte-unixbench/UnixBench/Makefile', inplace=True):
                if 'CC=gcc' in line:  # 仅在包含'CC=gcc'的行进行替换
                    line = line.replace('CC=gcc', 'CC=clang')
                print(line, end='') #inplace=True会把stdout重定向到文件中

        try:
            subprocess.run(
                "make",
                cwd="/root/osmts_tmp/byte-unixbench/UnixBench",
                shell=True,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.PIPE
            )
        except subprocess.CalledProcessError as e:
            raise CompileError(e.returncode,self.compiler,e.stderr)


    def run_test(self):
        try:
            run = subprocess.run(
                "./Run -c 1 -c $(nproc)",
                cwd="/root/osmts_tmp/byte-unixbench/UnixBench/",
                shell=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE
            )
        except subprocess.CalledProcessError as e:
            raise RunError(e.returncode,e.stderr)
        else:
            self.test_result = run.stdout.decode('utf-8')
            with open(self.directory / 'unixbench,log','w') as file:
                file.write(self.test_result)


    def result2summary(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "unixbench"
        ws['a1'] = "unixbench"
        ws['b1'] = "测试项目"
        ws['c1'] = "数据"

        ws['a2'] = "64 CPUs & 1 parallel"
        ws['a15'] = "64 CPUs & 64 parallel"
        ws.merge_cells("a2:a13")
        ws.merge_cells("a15:a26")

        ws['b2'] = ws['b15'] = "Dhrystone 2 using register variables(lps)"
        ws['b3'] = ws['b16'] = "Double-Precision Whetstone(MWIPS)"
        ws['b4'] = ws['b17'] = "Execl Throughput(lps)"
        ws['b5'] = ws['b18'] = "File Copy 1024 bufsize 2000 maxblocks(KBps)"
        ws['b6'] = ws['b19'] = "File Copy 256 bufsize 500 maxblocks(KBps)"
        ws['b7'] = ws['b20'] = "File Copy 4096 bufsize 8000 maxblocks(KBps)"
        ws['b8'] = ws['b21'] = "Pipe Throughput(lps)"
        ws['b9'] = ws['b22'] = "Pipe-based Context Switching(lps)"
        ws['b10'] = ws['b23'] = "Process Creation(lps)"
        ws['b11'] = ws['b24'] = "Shell Scripts (1 concurrent)(lpm)"
        ws['b12'] = ws['b25'] = "Shell Scripts (8 concurrent)(lpm)"
        ws['b13'] = ws['b26'] = "System Call Overhead(lps)"

        match_list = [
            re.compile(r"Dhrystone 2 using register variables\s+([\d.]+\d)\slps"),
            re.compile(r"Double-Precision Whetstone\s+([\d.]+\d)\sMWIPS"),
            re.compile(r"Execl Throughput\s+([\d.]+\d)\slps"),
            re.compile(r"File Copy 1024 bufsize 2000 maxblocks\s+([\d.]+\d)\sKBps"),
            re.compile(r"File Copy 256 bufsize 500 maxblocks\s+([\d.]+\d)\sKBps"),
            re.compile(r"File Copy 4096 bufsize 8000 maxblocks\s+([\d.]+\d)\sKBps"),
            re.compile(r"Pipe Throughput\s+([\d.]+\d)\slps"),
            re.compile(r"Pipe-based Context Switching\s+([\d.]+\d)\slps"),
            re.compile(r"Process Creation\s+([\d.]+\d)\slps"),
            re.compile(r"Shell Scripts \(1 concurrent\)\s+([\d.]+\d)\slpm"),
            re.compile(r"Shell Scripts \(8 concurrent\)\s+([\d.]+\d)\slpm"),
            re.compile(r"System Call Overhead\s+([\d.]+\d)\slps"),
        ]
        line = 2
        for one in match_list:
            match_result = one.findall(self.test_result)
            ws[f'c{line}'] = match_result[0]
            ws[f'c{line + 13}'] = match_result[1]
            line += 1

        wb.save(self.directory / 'unixbench.xlsx')


    def run(self):
        print("开始进行unixbench测试")
        self.pre_test()
        self.run_test()
        try:
            self.result2summary()
        except Exception as e:
            logFile = self.directory / 'unixbench_summary_error.log'
            with open(logFile, 'w') as log:
                log.write(str(e))
            raise SummaryError(logFile)
        print("unixbench测试结束")